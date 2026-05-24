"""
Exportador de dados do IRPF para planilha Excel (.xlsx) em PT-BR.

Abas geradas:
  - "Resumo"
  - "Bens e Direitos"
  - "Rendimentos Isentos"
  - "Rendimentos Exclusivos"
"""
from __future__ import annotations

import logging
from decimal import Decimal
from typing import List, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from project.models.canonical import AssetRecord, ExemptIncomeRecord, ExclusiveIncomeRecord, TaxableIncomeRecord

logger = logging.getLogger(__name__)

SUMMARY_SHEET = "Resumo"
ASSETS_SHEET = "Bens e Direitos"
EXEMPTS_SHEET = "Rendimentos Isentos"
EXCLUSIVES_SHEET = "Rendimentos Exclusivos"
TAXABLES_SHEET = "Rendimentos Tributáveis"


def _money(value: Decimal) -> float:
    return float(value or Decimal("0"))


def _decimal_or_blank(value: Decimal | None) -> float | str:
    return float(value) if value is not None else ""


def _auto_width(ws) -> None:
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def _style_header_row(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _make_table(ws, display_name: str) -> None:
    if ws.max_row < 2:
        return
    tab = Table(displayName=display_name, ref=ws.dimensions)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)
    ws.freeze_panes = ws["A2"]


def _hyperlink_formula(target_sheet: str, cell_ref: str, label: str) -> str:
    safe_label = label.replace('"', "'")
    return f'=HYPERLINK("#\'{target_sheet}\'!{cell_ref}","{safe_label}")'


def _build_asset_index(assets: Sequence[AssetRecord]) -> dict[str, AssetRecord]:
    index: dict[str, AssetRecord] = {}
    for asset in assets:
        if asset.cnpj_fonte:
            index[asset.cnpj_fonte] = asset
    return index


def _set_hyperlink(cell, formula: str) -> None:
    cell.value = formula
    cell.style = "Hyperlink"


def _item_label(item_number: int, base_label: str) -> str:
    return f"{base_label} item {item_number}"


def _write_summary(ws, assets: Sequence[AssetRecord], asset_rows: dict[int, int]) -> dict[int, int]:
    headers = [
        "Item",
        "Grupo",
        "Codigo",
        "Codigo do Bem",
        "Descricao do Bem",
        "Nome do Ativo",
        "Localizacao",
        "CNPJ Fonte",
        "Valor Ano Anterior (R$)",
        "Valor Ano Atual (R$)",
        "Total Rend. Isentos (R$)",
        "Total Rend. Exclusivos (R$)",
        "Total Rend. Tributáveis (R$)",
        "Ver Detalhes",
    ]
    ws.append(headers)

    summary_rows: dict[int, int] = {}
    for idx, asset in enumerate(assets, start=2):
        asset_row = asset_rows[id(asset)]
        summary_rows[id(asset)] = idx
        ws.append([
            idx - 1,
            asset.grupo_bem,
            asset.codigo_item,
            asset.codigo_bem,
            asset.descricao,
            asset.nome_ativo,
            asset.localizacao,
            asset.cnpj_fonte,
            _money(asset.valor_anterior),
            _money(asset.valor_2025),
            _money(sum((r.valor for r in asset.rendimentos_isentos), Decimal("0"))),
            _money(sum((r.valor for r in asset.rendimentos_exclusivos), Decimal("0"))),
            _money(sum((r.valor for r in asset.rendimentos_tributaveis), Decimal("0"))),
            None,
        ])
        _set_hyperlink(
            ws.cell(row=idx, column=14),
            _hyperlink_formula(ASSETS_SHEET, f"A{asset_row}", _item_label(idx - 1, "Abrir")),
        )

    _style_header_row(ws)
    _auto_width(ws)
    _make_table(ws, "TabelaResumo")
    return summary_rows


def _write_assets(ws, assets: Sequence[AssetRecord]) -> dict[int, int]:
    headers = [
        "Item",
        "CPF",
        "Grupo",
        "Codigo",
        "Codigo do Bem",
        "Descricao",
        "Localizacao",
        "Indicador Exterior",
        "Codigo do Pais",
        "CNPJ Fonte",
        "Discriminacao",
        "Situacao Anterior (R$)",
        "Valor Atual (R$)",
        "Rend. Isentos Vinculados",
        "Rend. Exclusivos Vinculados",
        "Rend. Tributáveis Vinculados",
        "Instituicao",
        "Nome do Ativo",
        "Quantidade",
        "Preco Medio",
    ]
    ws.append(headers)

    asset_rows: dict[int, int] = {}
    for asset in assets:
        row_idx = ws.max_row + 1
        asset_rows[id(asset)] = row_idx
        ws.append([
            row_idx - 1,
            asset.cpf,
            asset.grupo_bem,
            asset.codigo_item,
            asset.codigo_bem,
            asset.descricao,
            asset.localizacao,
            asset.indicador_exterior,
            asset.codigo_pais,
            asset.cnpj_fonte,
            asset.discriminacao,
            _money(asset.valor_anterior),
            _money(asset.valor_2025),
            len(asset.rendimentos_isentos),
            len(asset.rendimentos_exclusivos),
            len(asset.rendimentos_tributaveis),
            asset.instituicao,
            asset.nome_ativo,
            asset.quantidade if asset.quantidade is not None else "",
            _decimal_or_blank(asset.preco_medio),
        ])

    _style_header_row(ws)
    _auto_width(ws)
    _make_table(ws, "TabelaBens")
    return asset_rows


def _write_income_sheet(
    ws,
    records: Sequence[ExemptIncomeRecord | ExclusiveIncomeRecord | TaxableIncomeRecord],
    assets_by_cnpj: dict[str, AssetRecord],
    asset_rows: dict[int, int],
    *,
    display_name: str,
    bem_header: str = "Bem",
) -> dict[int, int]:
    headers = [
        "Item",
        "CPF",
        "Codigo",
        "Descricao",
        "Valor (R$)",
        "CNPJ Fonte Pagadora",
        "Nome Fonte Pagadora",
        bem_header,
        "Origem",
    ]
    ws.append(headers)

    first_rows: dict[int, int] = {}
    for record in records:
        asset = assets_by_cnpj.get(record.cnpj_fonte, None) if record.cnpj_fonte else None
        row_idx = ws.max_row + 1
        if asset and id(asset) not in first_rows:
            first_rows[id(asset)] = row_idx

        ws.append([
            row_idx - 1,
            record.cpf,
            record.tipo_rendimento,
            record.descricao,
            _money(record.valor),
            record.cnpj_fonte,
            record.nome_fonte,
            asset.descricao if asset else (record.bem_associado or ""),
            "Detalhe" if record.origem == "detalhe" else "Direto",
        ])

        if asset and id(asset) in asset_rows:
            _set_hyperlink(
                ws.cell(row=row_idx, column=8),
                _hyperlink_formula(
                    ASSETS_SHEET,
                    f"A{asset_rows[id(asset)]}",
                    _item_label(asset_rows[id(asset)] - 1, asset.descricao),
                ),
            )

    _style_header_row(ws)
    _auto_width(ws)
    _make_table(ws, display_name)
    return first_rows


def _write_asset_income_links(
    ws,
    assets: Sequence[AssetRecord],
    asset_rows: dict[int, int],
    exempt_rows: dict[int, int],
    exclusive_rows: dict[int, int],
) -> None:
    for asset in assets:
        row_idx = asset_rows[id(asset)]
        exempt_count = len(asset.rendimentos_isentos)
        exclusive_count = len(asset.rendimentos_exclusivos)

        exempt_cell = ws.cell(row=row_idx, column=14)
        if exempt_count and id(asset) in exempt_rows:
            _set_hyperlink(
                exempt_cell,
                _hyperlink_formula(
                    EXEMPTS_SHEET,
                    f"A{exempt_rows[id(asset)]}",
                    _item_label(exempt_rows[id(asset)] - 1, str(exempt_count)),
                ),
            )
        else:
            exempt_cell.value = exempt_count

        exclusive_cell = ws.cell(row=row_idx, column=15)
        if exclusive_count and id(asset) in exclusive_rows:
            _set_hyperlink(
                exclusive_cell,
                _hyperlink_formula(
                    EXCLUSIVES_SHEET,
                    f"A{exclusive_rows[id(asset)]}",
                    _item_label(exclusive_rows[id(asset)] - 1, str(exclusive_count)),
                ),
            )
        else:
            exclusive_cell.value = exclusive_count


def export_to_excel(
    assets: List[AssetRecord],
    exempts: List[ExemptIncomeRecord],
    exclusives: List[ExclusiveIncomeRecord],
    taxables: List[TaxableIncomeRecord],
    output_path,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    ws_assets = wb.create_sheet(ASSETS_SHEET)
    asset_rows = _write_assets(ws_assets, assets)
    assets_by_cnpj = _build_asset_index(assets)

    ws_summary = wb.create_sheet(SUMMARY_SHEET, 0)
    summary_rows = _write_summary(ws_summary, assets, asset_rows)

    ws_exempts = wb.create_sheet(EXEMPTS_SHEET)
    exempt_rows = _write_income_sheet(
        ws_exempts,
        exempts,
        assets_by_cnpj,
        asset_rows,
        display_name="TabelaRendIsentos",
    )

    ws_exclusives = wb.create_sheet(EXCLUSIVES_SHEET)
    exclusive_rows = _write_income_sheet(
        ws_exclusives,
        exclusives,
        assets_by_cnpj,
        asset_rows,
        display_name="TabelaRendExclusivos",
    )

    ws_taxables = wb.create_sheet(TAXABLES_SHEET)
    taxable_rows = _write_income_sheet(
        ws_taxables,
        taxables,
        assets_by_cnpj,
        asset_rows,
        display_name="TabelaRendTributaveis",
        bem_header="Bem Associado (Descrição)",
    )

    _write_asset_income_links(ws_assets, assets, asset_rows, exempt_rows, exclusive_rows)
    for asset in assets:
        row_idx = asset_rows[id(asset)]
        taxable_count = len(asset.rendimentos_tributaveis)
        taxable_cell = ws_assets.cell(row=row_idx, column=16)
        if taxable_count and id(asset) in taxable_rows:
            _set_hyperlink(
                taxable_cell,
                _hyperlink_formula(
                    TAXABLES_SHEET,
                    f"A{taxable_rows[id(asset)]}",
                    _item_label(taxable_rows[id(asset)] - 1, str(taxable_count)),
                ),
            )
        else:
            taxable_cell.value = taxable_count

    _auto_width(ws_assets)

    if not wb.sheetnames:
        wb.create_sheet("Sem Dados")

    try:
        wb.save(str(output_path))
        logger.info("Arquivo Excel gravado em: %s", output_path)
    except Exception as exc:
        logger.error("Falha ao gravar arquivo Excel: %s", exc)
        raise
